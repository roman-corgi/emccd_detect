# -*- coding: utf-8 -*-
"""Simulation for EMCCD detector."""

import os
import warnings
from pathlib import Path

import numpy as np
import scipy as sci
import openpyxl as xl

from emccd_detect.cosmics import cosmic_hits, sat_tails
from emccd_detect.rand_em_gain import rand_em_gain
from emccd_detect.util.read_metadata_wrapper import MetadataWrapper
from arcticpy import add_cti, CCD, ROE, Trap


class EMCCDDetectException(Exception):
    """Exception class for emccd_detect module."""


class EMCCDDetectBase:
    """Base class for EMCCD detector.

    Parameters
    ----------
    em_gain : float
        Electron multiplying gain (e-/photoelectron).
    full_well_image : float
        Image area full well capacity (e-).
    status:  {0,1}
        0 if you want to ignore the effect on the y-intercept in PTC curves.
        1 if you want to include the effect.
    dark_current: float
        Dark current rate (e-/pix/s).
    cic : float
        Clock induced charge (e-/pix/frame).
    read_noise : float
        Read noise (e-/pix/frame).
    bias : float
        Bias offset (e-).
    qe : float
        Quantum efficiency.
    cr_rate : float
        Cosmic ray rate (hits/cm^2/s).
    pixel_pitch : float
        Distance between pixel centers (m).
    eperdn : float
        Electrons per dn.
    nbits : int
        Number of bits used by the ADC readout. Must be between 1 and 64,
        inclusive.
    numel_gain_register : int
        Number of gain register elements. For modeling partial CIC.

    """
    def __init__(
        self,
        em_gain,
        full_well_image,
        #full_well_serial,
        status,
        dark_current,
        cic,
        read_noise,
        bias,
        qe,
        cr_rate,
        pixel_pitch,
        eperdn,
        nbits,
        numel_gain_register
    ):
        # Input checks
        if not isinstance(nbits, (int, np.integer)):
            raise EMCCDDetectException('nbits must be an integer')
        if nbits < 1 or nbits > 64:
            raise EMCCDDetectException('nbits must be between 1 and 64, '
                                       'inclusive')

        self.em_gain = em_gain
        self.full_well_image = full_well_image
        #self.full_well_serial = full_well_serial
        self.status = status
        self.dark_current = dark_current
        self.cic = cic
        self.read_noise = read_noise
        self.bias = bias
        self.qe = qe
        self.cr_rate = cr_rate
        self.pixel_pitch = pixel_pitch
        self.eperdn = eperdn
        self.nbits = nbits
        self.numel_gain_register = numel_gain_register

        # Placeholders for trap parameters
        self.ccd = None
        self.roe = None
        self.traps = None
        self.express = None
        self.offset = None
        self.window_range = None

        # Placeholders for derived values
        self.mean_expected_rate = None

    @property
    def eperdn(self):
        return self._eperdn

    @eperdn.setter
    def eperdn(self, eperdn):
        try:
            eperdn = float(eperdn)
        except Exception:
            raise EMCCDDetectException('eperdn value must be a float')

        if eperdn <= 0:
            raise EMCCDDetectException('eperdn value must be positve.')
        else:
            self._eperdn = eperdn

    def update_cti(
        self,
        ccd=None,
        roe=None,
        traps=None,
        express=1,
        offset=0,
        window_range=None
    ):
        # Update parameters
        self.ccd = ccd
        self.roe = roe
        self.traps = traps

        self.express = express
        self.offset = offset
        self.window_range = window_range

        # Instantiate defaults for any class instances not provided
        if self.ccd is None:
            self.ccd = CCD()
        if roe is None:
            self.roe = ROE()
        if traps is None:
            self.traps = [Trap()]

    def unset_cti(self):
        # Remove CTI simulation
        self.ccd = None
        self.roe = None
        self.traps = None

    def sim_sub_frame(self, fluxmap, frametime):
        """Simulate a partial detector frame.

        This runs the same algorithm as sim_full_frame, but only on the given
        fluxmap without surrounding it with prescan/overscan. The input fluxmap
        array may be arbitrary in shape and an image array of the same shape
        will be returned.

        Parameters
        ----------
        fluxmap : array_like
            Input fluxmap of arbitrary shape (phot/pix/s).
        frametime : float
            Frame exposure time (s).

        Returns
        -------
        output_counts : array_like
            Detector output counts, same shape as input fluxmap (dn).

        Notes
        -----
        This method is just as accurate and will return the same results as if
        the user ran sim_full_frame and then subsectioned the input fluxmap,
        with the exception of cosmic tails.

        It is slightly less accurate when cosmics are used, since the tail
        wrapping will be too strong. In a full frame the cosmic tails wrap into
        the next row in the prescan and trail off significantly before getting
        back to the image area, but here there is no prescan so the tails will
        be immediately wrapped back into the image.

        """
        # Simulate the integration process
        exposed_pix_m = np.ones_like(fluxmap).astype(bool)  # No unexposed pixels
        actualized_e = self.integrate(fluxmap.copy(), frametime, exposed_pix_m)

        # Simulate parallel clocking
        parallel_counts = self.clock_parallel(actualized_e)

        # Simulate serial clocking (output will be flattened to 1d)
        empty_element_m = np.zeros_like(parallel_counts).astype(bool)  # No empty elements
        gain_counts = self.clock_serial(parallel_counts, empty_element_m)

        # Simulate amplifier and adc redout
        output_dn = self.readout(gain_counts)

        # Reshape from 1d to 2d
        return output_dn.reshape(actualized_e.shape)

    def full_well_serial_func(self):
        """
        Returns fwc_em.  Dependence was derived from an interpolation.
        """
        if self.em_gain <=1.06082:  #corresponding to hv = 22 (1.0608 in fit function, but we know it to be 1 hv=22)
            return 100000
        if self.em_gain > 1.06082:
            return int(np.floor((6917529027641081856000*(3/2)**(337164395852765312/6243314768165359- (11258999068426240*np.log(76451918253118240*np.log(self.em_gain)))/6243314768165359))/1332894850849759 + 40000))

    def _slope(self):
        """
        Fitted function for slope as it differs from 0.5 (shot noise).
        """
            #fitSlope=  0.5 +2.^(hv/2.3)/2.^(22/2.3)/750;
            #fitG = 10.^(2.^(hv0/2.5)/2.^(22/2.5)/39);
            #invert the function to get gain(hv):
            #hv = -(210*log(2) + 5*log(log(10))...
            #    - 5*log(76451918253118239*log(g)))/(2*log(2));
        if self.em_gain <=1:
            return 0.5
        if self.em_gain > 1:
            return (4398046511104*2**((225179981368524800*np.log(76451918253118240*np.log(self.em_gain)))/143596239667803257 - 6743287917055306240/143596239667803257))/2498839895518397625 + 1/2

    def _ypow(self):
        """
        Fitted function for the power to which the factor representing the
        y-intercept of PTC curves should be raised.  Linear interpolation for
        the y-intercepts was good enough for this process.

        status:  {0, 1}
            0 if you don't want to use the y-intercept constraint (assume regular form in the noise), 1 if you want to use it.
        """
        if self.status == 0:
            return 0.5
        if self.status == 1:
            if self.em_gain <=1.495:  #g corresponding to hv=30 (exactly, not from fitted function, but doesn't matter too much either way
                return -0.0972
            if self.em_gain > 1.495:
                hv =  -(210*np.log(2) + 5*np.log(np.log(10))- 5*np.log(76451918253118239*np.log(self.em_gain)))/(2*np.log(2))
                b = (hv-30)*(0.1253 + 0.4836)/9 - 0.4836
                #return (b-(_slope(g)-1)*np.log10(8))/(np.log10((_ENF(g,604))**2*g))
                #replace _slope(g) with ((1.07882-.789911)/(719.766-1.495)*
                # (g-1.495)+0.789911), which essentially breaks up _slope(g)
                # into as many steps as there are in g between HV=30 and 39;
                # need to do this b/c I originally assumed linear fit for y-int,
                # and it really isn't linear
                return (b-(((1.07882-.789911)/(719.766-1.495)*(self.em_gain-1.495)+0.789911)-1)*np.log10(8))/(np.log10((self._ENF())**2*self.em_gain)) #specific to 604 EMCCD

    #wrong and not needed
    # def _z(self,counts):
    #     """_z is the effective slope after taking into account the difference
    #     in y-intercept on the PTC curves."""
    #     g = self.em_gain
    #     # counts is the number of counts after gain ("gain_counts")

    #     return self._ypow()*np.emath.logn(counts,g*self._ENF()**2)+self._slope()

    def _ENF(self):
        """
        Returns the ENF.
        """
        Nem = self.numel_gain_register
        g = self.em_gain
        return np.sqrt(2*(g-1)*g**(-(Nem +1)/Nem) + 1/g)



    def integrate(self, fluxmap_full, frametime, exposed_pix_m):
        # Add cosmic ray effects
        # XXX Maybe change this to units of flux later
        cosm_actualized_e = cosmic_hits(np.zeros_like(fluxmap_full),
                                        self.cr_rate, frametime,
                                        self.pixel_pitch, self.full_well_image)

        # Mask flux out of unexposed (covered) pixels
        fluxmap_full[~exposed_pix_m] = 0
        cosm_actualized_e[~exposed_pix_m] = 0

        # Simulate imaging area pixel effects over time
        actualized_e = self._imaging_area_elements(fluxmap_full, frametime,
                                                   cosm_actualized_e)

        return actualized_e

    def clock_parallel(self, actualized_e):
        # Only add CTI if update_cti has been called
        if self.ccd is not None and self.roe is not None and self.traps is not None:
            parallel_counts = add_cti(
                actualized_e.copy(),
                parallel_roe=self.roe,
                parallel_ccd=self.ccd,
                parallel_traps=self.traps,
                parallel_express=self.express,
                parallel_offset=self.offset,
                parallel_window_range=self.window_range
            )
        else:
            parallel_counts = actualized_e

        return parallel_counts

    def clock_serial(self, actualized_e_full, empty_element_m):
        # Actualize cic electrons in prescan and overscan pixels
        # XXX Another place where we are fudging a little
        actualized_e_full[empty_element_m] = np.random.poisson(actualized_e_full[empty_element_m]
                                                               + self.cic)
        # XXX Call arcticpy here
        # Flatten row by row
        actualized_e_full_flat = actualized_e_full.ravel()

        # Clock electrons through serial register elements
        serial_counts = self._serial_register_elements(actualized_e_full_flat)

        # Clock electrons through gain register elements
        gain_counts = self._gain_register_elements(serial_counts)

        return gain_counts

    def readout(self, gain_counts):
        # Pass electrons through amplifier
        amp_ev = self._amp(gain_counts)

        # Pass amp electron volt counts through analog to digital converter
        output_dn = self._adc(amp_ev)

        return output_dn

    def _imaging_area_elements(self, fluxmap_full, frametime, cosm_actualized_e):
        """Simulate imaging area pixel behavior for a given fluxmap and
        frametime.

        Note that the imaging area is defined as the active pixels which are
        exposed to light plus the surrounding dark reference and transition
        areas, which are covered and recieve no light. These pixels are
        indentical to the active area, so while they recieve none of the
        fluxmap they still have the same noise profile.

        Parameters
        ----------
        fluxmap_full : array_like
            Incident photon rate fluxmap (phot/pix/s).
        frametime : float
            Frame exposure time (s).
        cosm_actualized_e : array_like
            Electrons actualized from cosmic rays, same size as fluxmap_full (-e).

        Returns
        -------
        actualized_e : array_like
            Map of actualized electrons (e-).

        """
        # Calculate mean photo-electrons after integrating over frametime
        mean_phe_map = fluxmap_full * frametime * self.qe

        # Calculate mean expected rate after integrating over frametime
        mean_dark = self.dark_current * frametime
        mean_noise = mean_dark + self.cic

        # Set mean expected rate (commonly referred to as lambda)
        self.mean_expected_rate = mean_phe_map + mean_noise

        # Actualize electrons at the pixels
        actualized_e = np.random.poisson(self.mean_expected_rate).astype(float)

        # Add cosmic ray effects
        # XXX Maybe change this to units of flux later
        actualized_e += cosm_actualized_e

        # Cap at pixel full well capacity
        actualized_e[actualized_e > self.full_well_image] = self.full_well_image

        return actualized_e

    def _serial_register_elements(self, actualized_e_full_flat):
        """Simulate serial register element behavior.

        Parameters
        ----------
        actualized_e_full_flat : array_like
            Electrons actualized before clocking through the serial register.

        Returns
        -------
        serial_counts : array_like
            Electrons counts after passing through serial register elements.

        """
        serial_counts = actualized_e_full_flat
        return serial_counts

    def _gain_register_elements(self, serial_counts):
        """Simulate gain register element behavior and the degraded noise.

        Parameters
        ----------
        serial_counts : array_like
            Electrons counts after passing through serial register elements.

        Returns
        -------
        gain_counts : array_like
            Electron counts after passing through gain register elements.

        """
        # Apply EM gain
        full_well_serial = self.full_well_serial_func()
        gain_counts = np.zeros_like(serial_counts)
        gain_counts = rand_em_gain(
            n_in_array=serial_counts,
            em_gain=self.em_gain,
            max_out=full_well_serial
        )

        # Simulate saturation tails
        if self.cr_rate != 0:
            gain_counts = sat_tails(gain_counts, full_well_serial)

        # apply degraded noise (slope (s) and s')
        # mean is n*g
        g = self.em_gain
        mean = serial_counts*g
        deviation = gain_counts - mean
        #old and wrong:  enhfactor = gain_counts**(self._z(gain_counts)-0.5)
        F = self._ENF()
        enhfactor = (g*F**2)**(self._ypow()-.5)*mean**(self._slope()-.5)
        gain_counts_enh = deviation*enhfactor + mean

        # apply non-linearity
        # old and wrong: added gain_counts_enh*self.percent_NL(gain_counts_slope)/100
        gain_counts = gain_counts_enh + mean*self.percent_NL(mean)/100

        # Cap at full well capacity of gain register
        gain_counts[gain_counts > full_well_serial] = full_well_serial

        return gain_counts

    def percent_NL(self,counts):
        loc = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        #imported data from Excel spreadsheet from Nathan Bush
        nl_path = Path(loc, 'data', 'residual_non_linearity.xlsx')
        wb = xl.load_workbook(filename=nl_path)
        sheet = wb['Sheet1']
        # C5:C980 for nl_30, L5:L980 for nl_39
        percent_nl_30 = np.array([sheet.cell(i,3).value for i in range(5,981)])
        percent_nl_39 = np.array([sheet.cell(i,12).value for i in range(5,981)])
        sig = np.arange(500,98100,step=100)
        Nl_30 = sci.interpolate.PchipInterpolator(sig,percent_nl_30)
        Nl_39 = sci.interpolate.PchipInterpolator(sig,percent_nl_39)
        g = self.em_gain

        if g < 1.0608:  #corresponding to hv<22; hv 22 to 30 identical
            return Nl_30(counts)
        m = (Nl_39(counts) - Nl_30(counts))/(39-30)
        b = Nl_39(counts) - m*39
        hv = -(210*np.log(2) + 5*np.log(np.log(10))- 5*np.log(76451918253118239.*np.log(g)))/(2*np.log(2))
        if (g <= 719.7664) and ( g >= 1.0608): #between hv 30 and 39
            return m*hv+b
        if (g > 719.7664):  #hv above 39: use hv 39
            return Nl_39(counts)

    def _amp(self, serial_counts):
        """Simulate amp behavior.

        Parameters
        ----------
        serial_counts : array_like
            Electron counts from the serial register.

        Returns
        -------
        amp_ev : array_like
            Output from amp (eV).

        Notes
        -----
        Read noise is the amplifier read noise and not the effective read noise
        after the application of EM gain.

        """
        # Create read noise distribution in units of electrons
        read_noise_e = self.read_noise * np.random.normal(size=serial_counts.shape)

        # Apply read noise and bias to counts to get output electron volts
        amp_ev = serial_counts + read_noise_e + self.bias

        return amp_ev

    def _adc(self, amp_ev):
        """Simulate analog to digital converter behavior.

        Parameters
        ----------
        amp_ev : array_like
            Electron volt counts from amp (eV).

        Returns
        -------
        output_dn : array_like
            Analog to digital converter output (dn).

        """
        # Convert from electron volts to dn
        dn_min = 0
        dn_max = 2**self.nbits - 1
        output_dn = np.clip(amp_ev / self.eperdn, dn_min, dn_max).astype(np.uint64)

        return output_dn


class EMCCDDetect(EMCCDDetectBase):
    """Create an EMCCD-detected image for a given fluxmap.

    This class gives a method for simulating full frames (sim_full_frame) and
    also for adding simulated noise only to the input fluxmap (sim_sub_frame).

    Parameters
    ----------
    em_gain : float
        Electron multiplying gain (e-/photoelectron). Defaults to 5000.
    full_well_image : float
        Image area full well capacity (e-). Defaults to 60000.
    dark_current: float
        Dark current rate (e-/pix/s). Defaults to 0.0028.
    cic : float
        Clock induced charge (e-/pix/frame). Defaults to 0.02.
    read_noise : float
        Read noise (e-/pix/frame). Defaults to 100.
    bias : float
        Bias offset (e-). Defaults to 10000.
    qe : float
        Quantum efficiency. Defaults to 0.9.
    cr_rate : float
        Cosmic ray rate (hits/cm^2/s). Defaults to 0.
    pixel_pitch : float
        Distance between pixel centers (m). Defaults to 13e-6.
    eperdn : float
        Electrons per dn. Defaults to None.
    nbits : int
        Number of bits used by the ADC readout. Must be between 1 and 64,
        inclusive. Defaults to 14.
    numel_gain_register : int
        Number of gain register elements. For modeling partial CIC. Defaults to
        604.
    meta_path : str
        Full path of metadata yaml.

    """
    def __init__(
        self,
        em_gain=1.,
        full_well_image=60000.,
        #full_well_serial=None,
        status=1,
        dark_current=0.0028,
        cic=0.02,
        read_noise=100.,
        bias=10000.,
        qe=0.9,
        cr_rate=0.,
        pixel_pitch=13e-6,
        eperdn=None,
        nbits=14,
        numel_gain_register=604,
        meta_path=None
    ):
        # If no metadata file path specified, default to metadata.yaml in util
        if meta_path is None:
            here = os.path.abspath(os.path.dirname(__file__))
            meta_path = Path(here, 'util', 'metadata.yaml')

        # Before inheriting base class, get metadata
        self.meta_path = meta_path
        self.meta = MetadataWrapper(self.meta_path)

        # Set defaults from metadata
        #if full_well_serial is None:
        #    full_well_serial = self.meta.fwc
        if eperdn is None:
            eperdn = self.meta.eperdn

        super().__init__(
            em_gain=em_gain,
            full_well_image=full_well_image,
            #full_well_serial=full_well_serial,
            status=status,
            dark_current=dark_current,
            cic=cic,
            read_noise=read_noise,
            bias=bias,
            qe=qe,
            cr_rate=cr_rate,
            pixel_pitch=pixel_pitch,
            eperdn=eperdn,
            nbits=nbits,
            numel_gain_register=numel_gain_register
        )

    def sim_full_frame(self, fluxmap, frametime):
        """Simulate a full detector frame.

        Note that the fluxmap provided must be the same size as the exposed
        detector pixels (specified in self.meta.geom.image). A full frame
        including prescan and overscan regions will be made around the fluxmap.

        Parameters
        ----------
        fluxmap : array_like
            Input fluxmap of same shape as self.meta.geom.image (phot/pix/s).
        frametime : float
            Frame exposure time (s).

        Returns
        -------
        output_counts : array_like
            Detector output counts, including prescan/overscan (dn).

        """
        # Initialize the imaging area pixels
        imaging_area_zeros = self.meta.imaging_area_zeros.copy()
        # Embed the fluxmap within the imaging area. Create a mask for
        # referencing the input fluxmap subsection later
        fluxmap_full = self.meta.embed_im(imaging_area_zeros, 'image',
                                          fluxmap.copy())
        exposed_pix_m = self.meta.imaging_slice(self.meta.mask('image'))
        # Simulate the integration process
        actualized_e = self.integrate(fluxmap_full, frametime, exposed_pix_m)

        # Simulate parallel clocking
        parallel_counts = self.clock_parallel(actualized_e)

        # Initialize the serial register elements.
        full_frame_zeros = self.meta.full_frame_zeros.copy()
        # Embed the imaging area within the full frame. Create a mask for
        # referencing the prescan and overscan subsections later
        parallel_counts_full = self.meta.imaging_embed(full_frame_zeros, parallel_counts)
        empty_element_m = (self.meta.mask('prescan')
                           + self.meta.mask('parallel_overscan')
                           + self.meta.mask('serial_overscan'))
        # Simulate serial clocking
        gain_counts = self.clock_serial(parallel_counts_full, empty_element_m)

        # Simulate amplifier and adc redout
        output_dn = self.readout(gain_counts)

        # Reshape from 1d to 2d
        return output_dn.reshape(parallel_counts_full.shape)

    def slice_fluxmap(self, full_frame):
        """Return only the fluxmap portion of a full frame.

        Parameters
        ----------
        full_frame : array_like
            Simulated full frame.

        Returns
        -------
        array_like
            Fluxmap area of full frame.

        """
        return self.meta.slice_section(full_frame, 'image')

    def slice_prescan(self, full_frame):
        """Return only the prescan portion of a full frame.

        Parameters
        ----------
        full_frame : array_like
            Simulated full frame.

        Returns
        -------
        array_like
            Prescan area of a full frame.

        """
        return self.meta.slice_section(full_frame, 'prescan')

    def get_e_frame(self, frame_dn):
        """Take a raw frame output from EMCCDDetect and convert to a gain
        divided, bias subtracted frame in units of electrons.

        This will give the pre-readout image, i.e. the image in units of e- on
        the imaging plane.

        Parameters
        ----------
        frame_dn : array_like
            Raw output frame from EMCCDDetect, units of dn.

        Returns
        -------
        array_like
            Bias subtracted, gain divided frame in units of e-.

        """
        return (frame_dn * self.eperdn - self.bias) / self.em_gain


def emccd_detect(
    fluxmap,
    frametime,
    em_gain,
    full_well_image=50000.,
    #full_well_serial=90000.,
    status=1,
    dark_current=0.0028,
    cic=0.01,
    read_noise=100,
    bias=0.,
    qe=0.9,
    cr_rate=0.,
    pixel_pitch=13e-6,
    shot_noise_on=None
):
    """Create an EMCCD-detected image for a given fluxmap.

    This is a convenience function which wraps the base class implementation
    of the EMCCD simulator. It maintains the API of emccd_detect version 1.0.1.
    Note that output is in units of electrons, not dn.

    Parameters
    ----------
    fluxmap : array_like, float
        Input fluxmap (photons/pix/s).
    frametime : float
        Frame time (s).
    em_gain : float
        Electron multiplying gain (e-/photoelectron).
    full_well_image : float
        Image area full well capacity (e-). Defaults to 50000.
    dark_current: float
        Dark current rate (e-/pix/s). Defaults to 0.0028.
    cic : float
        Clock induced charge (e-/pix/frame). Defaults to 0.01.
    read_noise : float
        Read noise (e-/pix/frame). Defaults to 100.
    bias : float
        Bias offset (e-). Defaults to 0.
    qe : float
        Quantum efficiency. Defaults to 0.9.
    cr_rate : float
        Cosmic ray rate (hits/cm^2/s). Defaults to 0.
    pixel_pitch : float
        Distance between pixel centers (m). Defaults to 13e-6.
    shot_noise_on : bool, optional
        Apply shot noise. Defaults to None. [No longer supported as of v2.1.0.
        Input will have no effect.]

    Returns
    -------
    serial_frame : array_like, float
        Detector output (e-).

    Notes
    -----
    The value for eperdn (electrons per dn) is hardcoded to 1. This is for
    legacy purposes, as the version 1.0.1 implementation output electrons
    instead of dn.

    The legacy version also has no gain register CIC, so cic_gain_register is
    set to 0 and numel_gain_register is irrelevant.

    The legacy version also had no ADC (it just output floats), so the number
    of bits is set as high as possible (64) and the output is converted to
    floats. This will still be different from the legacy version as there will
    no longer be negative numbers.

    """
    if shot_noise_on is not None:
        warnings.warn('Shot noise parameter no longer supported. Input has no '
                      'effect')

    emccd = EMCCDDetectBase(
        em_gain=em_gain,
        full_well_image=full_well_image,
        #full_well_serial=full_well_serial,
        status=status,
        dark_current=dark_current,
        cic=cic,
        read_noise=read_noise,
        bias=bias,
        qe=qe,
        cr_rate=cr_rate,
        pixel_pitch=pixel_pitch,
        eperdn=1.,
        nbits=64,
        numel_gain_register=604
    )

    return emccd.sim_sub_frame(fluxmap, frametime).astype(float)
